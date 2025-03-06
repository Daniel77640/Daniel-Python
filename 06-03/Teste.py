import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook

# Lista de equipes para o menu suspenso
equipes = [
    "Sporting", "Porto", "Benfica", "Braga", "Maritimo", "Boavista", 
    "Vitória de Guimarães", "Rio Ave", "Famalicão", "Estoril Praia", 
    "Santa Clara", "Tondela", "Portimonense", "Arouca", "Casa Pia"
]

# Função para registrar os resultados no arquivo Excel
def registrar_resultado():
    time1 = time1_var.get()
    time2 = time2_var.get()
    gols_time1 = entry_gols_time1.get()
    gols_time2 = entry_gols_time2.get()
    data_jogo = entry_data_jogo.get()  # Data inserida pelo usuário
    
    # Verificando se todos os campos foram preenchidos
    if not time1 or not time2 or not gols_time1 or not gols_time2 or not data_jogo:
        messagebox.showerror("Erro", "Por favor, preencha todos os campos.")
        return
    
    try:
        gols_time1 = int(gols_time1)
        gols_time2 = int(gols_time2)
    except ValueError:
        messagebox.showerror("Erro", "Os gols devem ser números inteiros.")
        return

    # Validando a data
    try:
        # Validando a data inserida pelo usuário no formato DD-MM-YY
        from datetime import datetime
        datetime.strptime(data_jogo, '%d-%m-%y')
    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira a data no formato DD-MM-YY.")
        return

    # Determinando o vencedor
    if gols_time1 > gols_time2:
        vencedor = time1
    elif gols_time2 > gols_time1:
        vencedor = time2
    else:
        vencedor = "Empate"  # Caso os gols sejam iguais, é empate

    # Criando ou carregando o arquivo Excel
    try:
        wb = openpyxl.load_workbook('resultados_liga_portuguesa_2025.xlsx')
    except FileNotFoundError:
        # Se o arquivo não existir, cria um novo
        wb = Workbook()

    sheet = wb.active
    # Adicionando cabeçalhos, se necessário
    if sheet.max_row == 1:
        sheet.append(["Data", "Time 1", "Time 2", "Gols Time 1", "Gols Time 2", "Vencedor"])
    
    # Salvando os dados no arquivo
    sheet.append([data_jogo, time1, time2, gols_time1, gols_time2, vencedor])

    # Salvando o arquivo
    wb.save('resultados_liga_portuguesa_2025.xlsx')

    messagebox.showinfo("Sucesso", f"Resultado de {time1} vs {time2} registrado com sucesso!")

    # Limpar os campos
    entry_gols_time1.delete(0, tk.END)
    entry_gols_time2.delete(0, tk.END)
    entry_data_jogo.delete(0, tk.END)

# Configurando a interface gráfica com tkinter
root = tk.Tk()
root.title("Registro de Resultados - Liga Portuguesa 2025")

# Título
tk.Label(root, text="Registrar Resultado do Jogo", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10)

# Variáveis para os menus suspensos de equipes
time1_var = tk.StringVar()
time2_var = tk.StringVar()

# Menu suspenso para o time 1
tk.Label(root, text="Time 1:").grid(row=1, column=0, padx=10, pady=5)
time1_dropdown = tk.OptionMenu(root, time1_var, *equipes)
time1_dropdown.grid(row=1, column=1, padx=10, pady=5)

# Menu suspenso para o time 2
tk.Label(root, text="Time 2:").grid(row=2, column=0, padx=10, pady=5)
time2_dropdown = tk.OptionMenu(root, time2_var, *equipes)
time2_dropdown.grid(row=2, column=1, padx=10, pady=5)

# Entrada para gols do time 1
tk.Label(root, text="Gols Time 1:").grid(row=3, column=0, padx=10, pady=5)
entry_gols_time1 = tk.Entry(root)
entry_gols_time1.grid(row=3, column=1, padx=10, pady=5)

# Entrada para gols do time 2
tk.Label(root, text="Gols Time 2:").grid(row=4, column=0, padx=10, pady=5)
entry_gols_time2 = tk.Entry(root)
entry_gols_time2.grid(row=4, column=1, padx=10, pady=5)

# Entrada para data do jogo (apenas DD-MM-YY)
tk.Label(root, text="Data do Jogo (DD-MM-YY):").grid(row=5, column=0, padx=10, pady=5)
entry_data_jogo = tk.Entry(root)
entry_data_jogo.grid(row=5, column=1, padx=10, pady=5)

# Botão para registrar o resultado
button_registrar = tk.Button(root, text="Registrar Resultado", command=registrar_resultado)
button_registrar.grid(row=6, column=0, columnspan=2, pady=10)

# Iniciar a interface gráfica
root.mainloop()

