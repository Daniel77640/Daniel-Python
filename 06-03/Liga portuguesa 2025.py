import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook

# Lista de equipes para o menu suspenso
equipes = [
    "Sporting", "Porto", "Benfica", "Braga", "Maritimo", "Boavista", 
    "Vitória de Guimarães", "Rio Ave", "Famalicão", "Estoril Praia", 
    "Santa Clara", "Tondela", "Portimonense", "Arouca", "Casa Pia"
]

# Lista de jornadas
jornadas = [f"Jornada {i}" for i in range(1, 35)]  # Vamos ter 34 jornadas

# Função para registrar os resultados no arquivo Excel
def registrar_resultado():
    jornada = jornada_var.get()
    time1 = time1_var.get()
    time2 = time2_var.get()
    gols_time1 = entry_gols_time1.get()
    gols_time2 = entry_gols_time2.get()
    
    # Verificando se todos os campos foram preenchidos
    if not jornada or not time1 or not time2 or not gols_time1 or not gols_time2:
        messagebox.showerror("Erro", "Por favor, preencha todos os campos.")
        return
    
    try:
        gols_time1 = int(gols_time1)
        gols_time2 = int(gols_time2)
    except ValueError:
        messagebox.showerror("Erro", "Os gols devem ser números inteiros.")
        return

    # Determinando o vencedor e os pontos
    if gols_time1 > gols_time2:
        vencedor = time1
        pontos_time1 = 3
        pontos_time2 = 0
    elif gols_time2 > gols_time1:
        vencedor = time2
        pontos_time1 = 0
        pontos_time2 = 3
    else:
        vencedor = "Empate"
        pontos_time1 = 1
        pontos_time2 = 1

    # Criando ou carregando o arquivo Excel
    try:
        wb = openpyxl.load_workbook('resultados_liga_portuguesa_2025.xlsx')
    except FileNotFoundError:
        # Se o arquivo não existir, cria um novo
        wb = Workbook()

    sheet = wb.active
    # Adicionando cabeçalhos, se necessário
    if sheet.max_row == 1:
        header = ["Jornada", "Time 1", "Time 2", "Gols Time 1", "Gols Time 2", "Vencedor", "Pontos Time 1", "Pontos Time 2"]
        # Adiciona as colunas para o total de pontos de cada time
        header += [f"Total de Pontos {time}" for time in equipes]
        sheet.append(header)

    # Salvando os dados no arquivo
    sheet.append([jornada, time1, time2, gols_time1, gols_time2, vencedor, pontos_time1, pontos_time2])

    # Para cada equipe, somar os pontos acumulados (ajustando a fórmula)
    ultima_linha = sheet.max_row

    for i, time in enumerate(equipes, start=2):  # Começa da linha 2 (porque a 1ª é de cabeçalho)
        formula = f"SOMASE(B$2:B${ultima_linha}, \"{time}\", G$2:G${ultima_linha}) + SOMASE(C$2:C${ultima_linha}, \"{time}\", H$2:H${ultima_linha})"
        sheet[f"I{i}"] = formula  # Coloca a fórmula na coluna "I" para cada equipe

    # Salvando o arquivo
    wb.save('resultados_liga_portuguesa_2025.xlsx')

    messagebox.showinfo("Sucesso", f"Resultado de {time1} vs {time2} registrado com sucesso!")

    # Limpar os campos
    entry_gols_time1.delete(0, tk.END)
    entry_gols_time2.delete(0, tk.END)

# Configurando a interface gráfica com tkinter
root = tk.Tk()
root.title("Registro de Resultados - Liga Portuguesa 2025")

# Título
tk.Label(root, text="Registrar Resultado do Jogo", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10)

# Variáveis para os menus suspensos de equipes e jornada
jornada_var = tk.StringVar()
time1_var = tk.StringVar()
time2_var = tk.StringVar()

# Menu suspenso para a jornada
tk.Label(root, text="Jornada:").grid(row=1, column=0, padx=10, pady=5)
jornada_dropdown = tk.OptionMenu(root, jornada_var, *jornadas)
jornada_dropdown.grid(row=1, column=1, padx=10, pady=5)

# Menu suspenso para o time 1
tk.Label(root, text="Time 1:").grid(row=2, column=0, padx=10, pady=5)
time1_dropdown = tk.OptionMenu(root, time1_var, *equipes)
time1_dropdown.grid(row=2, column=1, padx=10, pady=5)

# Menu suspenso para o time 2
tk.Label(root, text="Time 2:").grid(row=3, column=0, padx=10, pady=5)
time2_dropdown = tk.OptionMenu(root, time2_var, *equipes)
time2_dropdown.grid(row=3, column=1, padx=10, pady=5)

# Entrada para gols do time 1
tk.Label(root, text="Gols Time 1:").grid(row=4, column=0, padx=10, pady=5)
entry_gols_time1 = tk.Entry(root)
entry_gols_time1.grid(row=4, column=1, padx=10, pady=5)

# Entrada para gols do time 2
tk.Label(root, text="Gols Time 2:").grid(row=5, column=0, padx=10, pady=5)
entry_gols_time2 = tk.Entry(root)
entry_gols_time2.grid(row=5, column=1, padx=10, pady=5)

# Botão para registrar o resultado
button_registrar = tk.Button(root, text="Registrar Resultado", command=registrar_resultado)
button_registrar.grid(row=6, column=0, columnspan=2, pady=10)

# Iniciar a interface gráfica
root.mainloop()

